from __future__ import annotations

import re
from collections import defaultdict
from dataclasses import dataclass
from datetime import date
from pathlib import Path
from typing import Iterable

from openpyxl import load_workbook

from .config import AppConfig
from .models import ActivityInterval, Issue
from .utils import normalize_text, parse_excel_date, parse_time_minutes


@dataclass
class SheetCandidate:
    file_path: str
    sheet_name: str
    start_date: object | None
    end_date: object | None
    agent_count: int
    row_count: int

    @property
    def label(self) -> str:
        period = "période inconnue"
        if self.start_date and self.end_date:
            period = f"{self.start_date.strftime('%d/%m/%Y')} → {self.end_date.strftime('%d/%m/%Y')}"
        return f"{Path(self.file_path).name} · {self.sheet_name} · {period} · {self.agent_count} agents"


class NiceWorkbookParser:
    AGENT_PATTERN = re.compile(r"^\s*agent\s*:\s*(\S+)\s+(.+?)\s*$", re.IGNORECASE)

    def __init__(self, config: AppConfig) -> None:
        self.config = config

    def find_candidates(self, file_paths: Iterable[str]) -> list[SheetCandidate]:
        candidates: list[SheetCandidate] = []
        for file_path in file_paths:
            workbook = load_workbook(file_path, read_only=True, data_only=True)
            try:
                for sheet in workbook.worksheets:
                    has_title = False
                    has_schedule_header = False
                    agents: set[str] = set()
                    dates = []
                    for row_index, row in enumerate(sheet.iter_rows(values_only=True), start=1):
                        values = list(row)
                        normalized = [normalize_text(value) for value in values if value not in (None, "")]
                        # NICE commonly writes the title with a curly apostrophe:
                        # "Horaires d’agent". During ASCII normalization, that apostrophe
                        # disappears and produces "horaires dagent" rather than
                        # "horaires d agent". Compare a compact representation so both
                        # curly and straight apostrophes, as well as spacing variants, work.
                        if any(self._is_schedule_title(value) for value in normalized):
                            has_title = True
                        if "date" in normalized and "activite planifiee" in normalized:
                            has_schedule_header = True
                        for value in values:
                            if isinstance(value, str):
                                match = self.AGENT_PATTERN.match(value)
                                if match:
                                    agents.add(match.group(1))
                        if row_index <= 40:
                            for value in values:
                                parsed = parse_excel_date(value)
                                if parsed:
                                    dates.append(parsed)
                        if row_index > 40 and has_title and agents:
                            # Enough evidence; dates later are recovered by full parsing.
                            pass
                    # The structural header is a safe fallback for exports whose title
                    # was renamed or omitted. Requiring an agent row prevents ordinary
                    # planning/output workbooks from being mistaken for a NICE extraction.
                    if agents and (has_title or has_schedule_header):
                        candidates.append(
                            SheetCandidate(
                                file_path=str(file_path),
                                sheet_name=sheet.title,
                                start_date=min(dates) if dates else None,
                                end_date=max(dates) if dates else None,
                                agent_count=len(agents),
                                row_count=sheet.max_row,
                            )
                        )
            finally:
                workbook.close()
        return candidates


    @staticmethod
    def _is_schedule_title(value: object) -> bool:
        compact = normalize_text(value).replace(" ", "")
        return compact in {
            "horairesdagent",
            "horairedagent",
            "horairesdesagents",
            "horairedesagents",
        }

    def choose_latest_per_file(self, candidates: list[SheetCandidate]) -> list[SheetCandidate]:
        grouped: dict[str, list[SheetCandidate]] = defaultdict(list)
        for candidate in candidates:
            grouped[candidate.file_path].append(candidate)
        selected = []
        for file_path, items in grouped.items():
            selected.append(
                max(
                    items,
                    key=lambda item: (
                        item.end_date or item.start_date or date.min,
                        item.start_date or date.min,
                        item.row_count,
                    ),
                )
            )
        return selected

    def parse(self, candidates: Iterable[SheetCandidate]) -> tuple[list[ActivityInterval], list[Issue]]:
        intervals: list[ActivityInterval] = []
        issues: list[Issue] = []
        by_file: dict[str, list[SheetCandidate]] = defaultdict(list)
        for candidate in candidates:
            by_file[candidate.file_path].append(candidate)

        for file_path, file_candidates in by_file.items():
            workbook = load_workbook(file_path, read_only=True, data_only=True)
            try:
                for candidate in file_candidates:
                    sheet = workbook[candidate.sheet_name]
                    parsed, sheet_issues = self._parse_sheet(file_path, sheet)
                    intervals.extend(parsed)
                    issues.extend(sheet_issues)
            finally:
                workbook.close()

        unique = {}
        for interval in intervals:
            key = (
                interval.agent_id,
                interval.work_date,
                interval.activity,
                interval.activity_start,
                interval.activity_end,
            )
            unique[key] = interval
        return sorted(unique.values(), key=lambda item: (item.work_date, item.agent_name, item.activity_start or -1)), issues

    def _parse_sheet(self, file_path: str, sheet) -> tuple[list[ActivityInterval], list[Issue]]:
        intervals: list[ActivityInterval] = []
        issues: list[Issue] = []
        current_agent_id = ""
        current_agent_name = ""
        current_date = None
        global_start = None
        global_end = None
        header = None
        pending = None

        known_non_eligible = {
            "pause repas",
            "libre",
            "repos",
            "conge paye",
            "form continue",
            "reunion equipe",
            "alternance ecole wh",
            "schedule modification",
            "abmtm",
        }

        def flush_pending(inferred_end=None) -> None:
            nonlocal pending
            if not pending:
                return
            activity_start = pending["activity_start"]
            explicit_end = pending["activity_end"]
            activity_end = explicit_end
            if activity_end is None or activity_end <= activity_start:
                activity_end = inferred_end
            if activity_end is None or activity_end <= activity_start:
                activity_end = pending["global_end"]
            if activity_end is None or activity_end <= activity_start:
                issues.append(
                    Issue(
                        severity="Avertissement",
                        issue_type="Ligne ignorée",
                        message=f"Fin d'activité introuvable : {pending['activity']!s}",
                        work_date=pending["work_date"],
                        agent_id=pending["agent_id"],
                        agent_name=pending["agent_name"],
                    )
                )
                pending = None
                return

            category = self.config.classify_activity(str(pending["activity"]))
            intervals.append(
                ActivityInterval(
                    source_file=str(file_path),
                    source_sheet=sheet.title,
                    source_row=pending["source_row"],
                    agent_id=pending["agent_id"],
                    agent_name=pending["agent_name"],
                    work_date=pending["work_date"],
                    global_start=pending["global_start"],
                    global_end=pending["global_end"],
                    activity=str(pending["activity"]).strip(),
                    activity_start=activity_start,
                    activity_end=activity_end,
                    category=category,
                )
            )
            if category.value == "Non éligible" and normalize_text(pending["activity"]) not in known_non_eligible:
                issues.append(
                    Issue(
                        severity="Information",
                        issue_type="Activité inconnue",
                        message=f"Activité non utilisée pour l'assistance : {pending['activity']}",
                        work_date=pending["work_date"],
                        agent_id=pending["agent_id"],
                        agent_name=pending["agent_name"],
                    )
                )
            pending = None

        for row_index, row in enumerate(sheet.iter_rows(values_only=True), start=1):
            values = list(row)
            agent_match = None
            for value in values:
                if isinstance(value, str):
                    match = self.AGENT_PATTERN.match(value)
                    if match:
                        agent_match = match
                        break
            if agent_match:
                flush_pending(global_end)
                current_agent_id = agent_match.group(1).strip()
                current_agent_name = agent_match.group(2).strip()
                current_date = None
                global_start = None
                global_end = None
                header = None
                continue

            normalized = [normalize_text(value) for value in values]
            if "date" in normalized and "activite planifiee" in normalized:
                flush_pending(global_end)
                header = self._detect_header(normalized)
                continue

            if not current_agent_id or not header:
                continue

            date_value = self._value(values, header["date"])
            parsed_date = parse_excel_date(date_value)
            if parsed_date:
                flush_pending(global_end)
                current_date = parsed_date
                global_start = parse_time_minutes(self._value(values, header["global_start"]))
                global_end = parse_time_minutes(self._value(values, header["global_end"]))
                first_value = normalize_text(self._value(values, header["global_start"]))
                if first_value in {"libre", "repos"}:
                    intervals.append(
                        ActivityInterval(
                            source_file=str(file_path),
                            source_sheet=sheet.title,
                            source_row=row_index,
                            agent_id=current_agent_id,
                            agent_name=current_agent_name,
                            work_date=current_date,
                            global_start=None,
                            global_end=None,
                            activity="Libre",
                            activity_start=None,
                            activity_end=None,
                        )
                    )
                    continue

            if not current_date:
                continue

            activity = self._activity_value(values, header)
            if not activity:
                continue
            activity_start = parse_time_minutes(self._value(values, header["activity_start"]))
            if activity_start is None:
                issues.append(
                    Issue(
                        severity="Avertissement",
                        issue_type="Ligne ignorée",
                        message=f"Début d'activité invalide : {activity!s}",
                        work_date=current_date,
                        agent_id=current_agent_id,
                        agent_name=current_agent_name,
                    )
                )
                continue

            # NICE sometimes omits the activity-end column. The next activity start
            # then closes the previous interval; the last interval ends at shift end.
            flush_pending(activity_start)
            pending = {
                "source_row": row_index,
                "agent_id": current_agent_id,
                "agent_name": current_agent_name,
                "work_date": current_date,
                "global_start": global_start,
                "global_end": global_end,
                "activity": activity,
                "activity_start": activity_start,
                "activity_end": self._activity_end(values, header),
            }

        flush_pending(global_end)
        return intervals, issues

    @staticmethod
    def _value(values, index):
        return values[index] if index is not None and index < len(values) else None

    @staticmethod
    def _detect_header(normalized: list[str]) -> dict[str, int | None]:
        date_index = normalized.index("date")
        activity_header = normalized.index("activite planifiee")
        starts = [index for index, value in enumerate(normalized) if value == "debut"]
        ends = [index for index, value in enumerate(normalized) if value == "fin"]
        global_start = starts[0] if starts else date_index + 1
        activity_start = starts[-1] if starts else activity_header + 2
        global_end = next((index for index in ends if index < activity_header), date_index + 2)
        activity_end = next((index for index in reversed(ends) if index > activity_start), None)
        return {
            "date": date_index,
            "global_start": global_start,
            "global_end": global_end,
            "activity_header": activity_header,
            "activity_start": activity_start,
            "activity_end": activity_end,
        }

    @staticmethod
    def _activity_value(values: list, header: dict[str, int]):
        left = header["global_end"] + 1
        right = header["activity_start"]
        for index in range(right - 1, left - 1, -1):
            if index < len(values) and values[index] not in (None, ""):
                return values[index]
        # NICE exports often place the value one cell to the right of the merged header.
        fallback = header["activity_header"] + 1
        return values[fallback] if fallback < len(values) else None

    @staticmethod
    def _activity_end(values: list, header: dict[str, int | None]):
        preferred = header.get("activity_end")
        if preferred is None or preferred >= len(values):
            return None
        return parse_time_minutes(values[preferred])
