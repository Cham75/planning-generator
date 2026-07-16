from __future__ import annotations

from collections import Counter, defaultdict
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

from .models import Category, ScheduleResult
from .utils import format_minutes, normalize_text


# Palette volontairement proche de l'ancien planning partagé.
DAY_BLUE = "D9EAF7"
HEADER_BLUE = "DCE6F1"
ASSIST_BLUE = "BDD7EE"
PAUSE_GREEN = "C6E0B4"
CALL_YELLOW = "FFF2CC"
ACTIVITY_ORANGE = "F8CBAD"
ABSENCE_ORANGE = "F4B183"
UNAVAILABLE_GRAY = "D9D9D9"
LIGHT_GRAY = "F2F2F2"
MISSING_RED = "FF0000"
WARNING_RED = "F4CCCC"
WHITE = "FFFFFF"
DARK = "1F1F1F"
BORDER_COLOR = "B7B7B7"
THIN = Side(style="thin", color=BORDER_COLOR)
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

FRENCH_DAYS = [
    "Lundi",
    "Mardi",
    "Mercredi",
    "Jeudi",
    "Vendredi",
    "Samedi",
    "Dimanche",
]

ROLE_ORDER = {
    "SUP": 0,
    "CQ": 1,
    "MO": 2,
    "TC": 3,
    "Team Lead": 4,
    "Open Time": 5,
    "Autre": 6,
    "À définir": 7,
}


class ExcelExporter:
    """Create a shared operational workbook with the daily planning as first sheet."""

    def export(self, result: ScheduleResult, output_path: str) -> None:
        workbook = Workbook()
        workbook.remove(workbook.active)
        self._planning_sheet(workbook, result)
        self._recap_sheet(workbook, result)
        self._controls_sheet(workbook, result)
        self._clean_data_sheet(workbook, result)
        workbook.active = 0
        workbook.save(output_path)

    def _planning_sheet(self, workbook: Workbook, result: ScheduleResult) -> None:
        sheet = workbook.create_sheet("Planning")
        sheet.sheet_view.showGridLines = False
        sheet.freeze_panes = "E3"

        slot_columns = {slot.key: index for index, slot in enumerate(result.slots, start=5)}
        assignment_keys = {
            (item.work_date, item.agent_id, item.slot_key)
            for item in result.assignments
        }
        assignments_by_day_agent = Counter(
            (item.work_date, item.agent_id) for item in result.assignments
        )
        days_by_date = {day.work_date: day for day in result.days}
        agent_days_by_date: dict = defaultdict(list)
        for agent_day in result.agent_days:
            agent_days_by_date[agent_day.work_date].append(agent_day)

        current_row = 1
        for work_date in sorted(days_by_date):
            day_schedule = days_by_date[work_date]
            agents = [
                agent
                for agent in agent_days_by_date[work_date]
                if self._show_agent_on_day(agent)
            ]
            agents.sort(
                key=lambda agent: (
                    ROLE_ORDER.get(agent.role, 99),
                    normalize_text(agent.name),
                )
            )

            # Day title, identical principle to the previous workbook.
            sheet.merge_cells(
                start_row=current_row,
                start_column=1,
                end_row=current_row,
                end_column=4 + len(result.slots),
            )
            title = sheet.cell(
                current_row,
                1,
                f"{FRENCH_DAYS[work_date.weekday()]} {work_date.strftime('%d/%m/%y')}",
            )
            title.font = Font(size=14, bold=True, color=DARK)
            title.fill = PatternFill("solid", fgColor=DAY_BLUE)
            title.alignment = Alignment(horizontal="left", vertical="center")
            sheet.row_dimensions[current_row].height = 24
            current_row += 1

            headers = ["Rôle", "Nom, Prénom", "Créneau", "# Assist. (h)"] + [
                slot.label for slot in result.slots
            ]
            for column, value in enumerate(headers, start=1):
                cell = sheet.cell(current_row, column, value)
                cell.font = Font(bold=True, color=DARK)
                cell.fill = PatternFill("solid", fgColor=HEADER_BLUE)
                cell.alignment = Alignment(
                    horizontal="center", vertical="center", wrap_text=True
                )
                cell.border = BORDER
            sheet.row_dimensions[current_row].height = 24
            current_row += 1

            for agent_day in agents:
                row = current_row
                assist_total = assignments_by_day_agent[(work_date, agent_day.agent_id)]
                sheet.cell(row, 1, agent_day.role)
                sheet.cell(row, 2, agent_day.name)
                sheet.cell(row, 3, self._working_window(agent_day))
                quota_cell = sheet.cell(
                    row,
                    4,
                    "Exclu" if agent_day.excluded else assist_total,
                )

                for column in range(1, 5):
                    cell = sheet.cell(row, column)
                    cell.border = BORDER
                    cell.alignment = Alignment(
                        horizontal="center" if column != 2 else "left",
                        vertical="center",
                        wrap_text=True,
                    )
                if agent_day.excluded:
                    quota_cell.fill = PatternFill("solid", fgColor=LIGHT_GRAY)
                    quota_cell.font = Font(italic=True, color="666666")
                elif agent_day.mandatory and assist_total == 0:
                    quota_cell.value = "0 / 1 min"
                    quota_cell.fill = PatternFill("solid", fgColor=WARNING_RED)
                    quota_cell.font = Font(bold=True, color="9C0006")

                for slot in result.slots:
                    column = slot_columns[slot.key]
                    assigned = (
                        work_date,
                        agent_day.agent_id,
                        slot.key,
                    ) in assignment_keys
                    value, fill, font_color = self._planning_slot_state(
                        agent_day,
                        slot.start_minute,
                        slot.end_minute,
                        assigned,
                    )
                    cell = sheet.cell(row, column, value)
                    cell.fill = PatternFill("solid", fgColor=fill)
                    cell.font = Font(
                        size=9,
                        bold=value == "Assistance",
                        color=font_color,
                    )
                    cell.alignment = Alignment(
                        horizontal="center",
                        vertical="center",
                        wrap_text=True,
                    )
                    cell.border = BORDER

                sheet.row_dimensions[row].height = 22
                current_row += 1

            # Coverage gap line directly below each daily team planning.
            gap_row = current_row
            sheet.merge_cells(
                start_row=gap_row, start_column=1, end_row=gap_row, end_column=4
            )
            gap_label = sheet.cell(gap_row, 1, "Manques assistance")
            gap_label.font = Font(bold=True)
            gap_label.alignment = Alignment(horizontal="left", vertical="center")
            gap_label.border = BORDER
            for col in range(2, 5):
                sheet.cell(gap_row, col).border = BORDER

            for slot in result.slots:
                missing = day_schedule.uncovered.get(slot.key, 0)
                cell = sheet.cell(
                    gap_row,
                    slot_columns[slot.key],
                    missing if missing else "",
                )
                cell.border = BORDER
                cell.alignment = Alignment(horizontal="center", vertical="center")
                if missing:
                    cell.fill = PatternFill("solid", fgColor=MISSING_RED)
                    cell.font = Font(bold=True, color=WHITE)
                else:
                    cell.fill = PatternFill("solid", fgColor=WHITE)
            sheet.row_dimensions[gap_row].height = 20
            current_row += 2

        # Shared-team layout and printing.
        widths = [11, 31, 17, 14] + [13] * len(result.slots)
        for index, width in enumerate(widths, start=1):
            sheet.column_dimensions[get_column_letter(index)].width = width
        sheet.page_setup.orientation = "landscape"
        sheet.page_setup.fitToWidth = 1
        sheet.page_setup.fitToHeight = 0
        sheet.sheet_properties.pageSetUpPr.fitToPage = True
        sheet.print_options.horizontalCentered = True
        sheet.print_area = f"A1:{get_column_letter(4 + len(result.slots))}{max(1, current_row - 1)}"

    def _recap_sheet(self, workbook: Workbook, result: ScheduleResult) -> None:
        sheet = workbook.create_sheet("Récap Assistances")
        dates = sorted(day.work_date for day in result.days)
        headers = [
            "Agent ID",
            "Nom, Prénom",
            "Rôle",
            "Exclu",
            "Total assistance (h)",
            "Jours avec assistance",
            "Heures consécutives",
        ] + [f"{FRENCH_DAYS[item.weekday()]} {item.strftime('%d/%m')}" for item in dates]
        sheet.append(headers)
        self._header(sheet[1])

        agent_days: dict[str, list] = defaultdict(list)
        assignments: dict[str, list] = defaultdict(list)
        for item in result.agent_days:
            agent_days[item.agent_id].append(item)
        for item in result.assignments:
            assignments[item.agent_id].append(item)
        slot_order = {slot.key: index for index, slot in enumerate(result.slots)}

        for agent_id, days in sorted(
            agent_days.items(), key=lambda item: normalize_text(item[1][0].name)
        ):
            items = assignments[agent_id]
            by_day = Counter(item.work_date for item in items)
            indexes_by_day: dict = defaultdict(list)
            for item in items:
                indexes_by_day[item.work_date].append(slot_order[item.slot_key])
            consecutive = sum(
                1
                for indexes in indexes_by_day.values()
                for first, second in zip(sorted(indexes), sorted(indexes)[1:])
                if second == first + 1
            )
            profile = days[0]
            row = [
                agent_id,
                profile.name,
                profile.role,
                "Oui" if all(day.excluded for day in days) else "Non",
                len(items),
                len(by_day),
                consecutive,
            ] + [by_day[item] for item in dates]
            sheet.append(row)

        for row in range(2, sheet.max_row + 1):
            for cell in sheet[row]:
                cell.border = BORDER
                cell.alignment = Alignment(vertical="center", wrap_text=True)
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = f"A1:{get_column_letter(sheet.max_column)}{sheet.max_row}"
        widths = [13, 31, 13, 10, 20, 21, 20] + [16] * len(dates)
        for index, width in enumerate(widths, start=1):
            sheet.column_dimensions[get_column_letter(index)].width = width
        sheet.sheet_view.showGridLines = False

    def _controls_sheet(self, workbook: Workbook, result: ScheduleResult) -> None:
        sheet = workbook.create_sheet("Contrôles")
        headers = ["Sévérité", "Date", "Type", "Agent ID", "Agent", "Message"]
        sheet.append(headers)
        self._header(sheet[1])
        if not result.issues:
            sheet.append(
                [
                    "OK",
                    None,
                    "Aucune anomalie",
                    "",
                    "",
                    "Toutes les règles contrôlées sont respectées.",
                ]
            )
        else:
            for issue in result.issues:
                sheet.append(
                    [
                        issue.severity,
                        issue.work_date,
                        issue.issue_type,
                        issue.agent_id,
                        issue.agent_name,
                        issue.message,
                    ]
                )
        for row in range(2, sheet.max_row + 1):
            severity = str(sheet.cell(row, 1).value)
            fill = (
                WARNING_RED
                if severity == "Erreur"
                else CALL_YELLOW
                if severity == "Avertissement"
                else PAUSE_GREEN
                if severity == "OK"
                else DAY_BLUE
            )
            for cell in sheet[row]:
                cell.border = BORDER
                cell.alignment = Alignment(vertical="top", wrap_text=True)
            sheet.cell(row, 1).fill = PatternFill("solid", fgColor=fill)
            if sheet.cell(row, 2).value:
                sheet.cell(row, 2).number_format = "dd/mm/yyyy"
        sheet.freeze_panes = "A2"
        self._add_table(sheet, "A1", f"F{sheet.max_row}", "ControlesTable")
        for index, width in enumerate([14, 13, 28, 13, 31, 90], start=1):
            sheet.column_dimensions[get_column_letter(index)].width = width
        sheet.sheet_view.showGridLines = False

    def _clean_data_sheet(self, workbook: Workbook, result: ScheduleResult) -> None:
        sheet = workbook.create_sheet("Données importées")
        headers = [
            "Fichier",
            "Feuille",
            "Ligne",
            "Agent ID",
            "Agent",
            "Date",
            "Début global",
            "Fin globale",
            "Activité",
            "Catégorie",
            "Début activité",
            "Fin activité",
        ]
        sheet.append(headers)
        self._header(sheet[1])
        for item in result.intervals:
            sheet.append(
                [
                    Path(item.source_file).name,
                    item.source_sheet,
                    item.source_row,
                    item.agent_id,
                    item.agent_name,
                    item.work_date,
                    format_minutes(item.global_start),
                    format_minutes(item.global_end),
                    item.activity,
                    item.category.value,
                    format_minutes(item.activity_start),
                    format_minutes(item.activity_end),
                ]
            )
        for row in range(2, sheet.max_row + 1):
            for cell in sheet[row]:
                cell.border = BORDER
                cell.alignment = Alignment(vertical="center", wrap_text=True)
            sheet.cell(row, 6).number_format = "dd/mm/yyyy"
        sheet.freeze_panes = "A2"
        self._add_table(
            sheet,
            "A1",
            f"L{sheet.max_row}",
            "DonneesImporteesTable",
        )
        widths = [28, 22, 9, 13, 31, 13, 14, 14, 28, 25, 16, 16]
        for index, width in enumerate(widths, start=1):
            sheet.column_dimensions[get_column_letter(index)].width = width
        sheet.sheet_view.showGridLines = False
        # Technical audit data remains available without distracting daily users.
        sheet.sheet_state = "hidden"

    @staticmethod
    def _header(cells) -> None:
        for cell in cells:
            cell.font = Font(bold=True, color=DARK)
            cell.fill = PatternFill("solid", fgColor=HEADER_BLUE)
            cell.alignment = Alignment(
                horizontal="center", vertical="center", wrap_text=True
            )
            cell.border = BORDER

    @staticmethod
    def _show_agent_on_day(agent_day) -> bool:
        if agent_day.global_start is not None or agent_day.global_end is not None:
            return True
        return any(
            normalize_text(interval.activity) not in {"libre", "repos"}
            for interval in agent_day.intervals
        )

    @staticmethod
    def _working_window(agent_day) -> str:
        if agent_day.global_start is None and agent_day.global_end is None:
            return "Libre"
        return f"{format_minutes(agent_day.global_start)}–{format_minutes(agent_day.global_end)}"

    @classmethod
    def _planning_slot_state(
        cls,
        agent_day,
        start: int,
        end: int,
        assigned: bool,
    ) -> tuple[str, str, str]:
        if assigned:
            return "Assistance", ASSIST_BLUE, DARK

        # Outside the global working window is visually greyed out.
        if (
            agent_day.global_start is None
            or agent_day.global_end is None
            or agent_day.global_start > start
            or agent_day.global_end < end
        ):
            return "", UNAVAILABLE_GRAY, DARK

        overlaps = [
            interval
            for interval in agent_day.intervals
            if interval.activity_start is not None
            and interval.activity_end is not None
            and interval.activity_start < end
            and interval.activity_end > start
        ]

        for interval in overlaps:
            if normalize_text(interval.activity) == "pause repas":
                return "Pause déjeuner", PAUSE_GREEN, DARK

        for interval in overlaps:
            if interval.category != Category.NON_ELIGIBLE:
                continue
            label, fill = cls._activity_label(interval.activity)
            return label, fill, DARK

        covering_eligible = [
            interval
            for interval in overlaps
            if interval.category != Category.NON_ELIGIBLE
            and interval.activity_start <= start
            and interval.activity_end >= end
        ]
        if covering_eligible:
            category = covering_eligible[0].category
            if start >= 18 * 60 and category in {
                Category.OPEN_TIME,
                Category.TEAM_LEAD,
            }:
                return "Prise d'appels", CALL_YELLOW, DARK
            if agent_day.excluded:
                return "", LIGHT_GRAY, DARK
            return "", WHITE, DARK

        # Inside the declared shift but not fully covered by a usable activity.
        return "", UNAVAILABLE_GRAY, DARK

    @staticmethod
    def _activity_label(activity: str) -> tuple[str, str]:
        normalized = normalize_text(activity)
        if "conge" in normalized or "rtt" in normalized:
            return "Congé", ABSENCE_ORANGE
        if any(word in normalized for word in ("absence", "abmtm", "maladie", "arret")):
            return "Absence", ABSENCE_ORANGE
        if "form" in normalized:
            return "Formation", ACTIVITY_ORANGE
        if "reunion" in normalized:
            return "Réunion", ACTIVITY_ORANGE
        if "alternance" in normalized or "ecole" in normalized:
            return "École / WH", ACTIVITY_ORANGE
        if normalized in {"libre", "repos"}:
            return "Repos", UNAVAILABLE_GRAY
        text = str(activity).strip()
        if len(text) > 18:
            text = text[:17] + "…"
        return text or "Indisponible", ACTIVITY_ORANGE

    @staticmethod
    def _add_table(sheet, start: str, end: str, name: str) -> None:
        table = Table(displayName=name, ref=f"{start}:{end}")
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        sheet.add_table(table)
