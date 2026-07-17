from __future__ import annotations

import tempfile
from collections import Counter, defaultdict
from datetime import date, timedelta
from pathlib import Path

from openpyxl import load_workbook

from planning_tool.config import AppConfig
from planning_tool.exporter import ExcelExporter
from planning_tool.models import ActivityInterval, AgentProfile, Category
from planning_tool.parser import NiceWorkbookParser
from planning_tool.scheduler import PlanningScheduler
from web_store import (
    BrowserLocalStorageAgentRepository,
    MemoryAgentStore,
    load_seed_profiles,
    profiles_from_json_bytes,
    profiles_to_json,
)

ROOT = Path(__file__).resolve().parent


class FakeBrowserStorage:
    def __init__(self) -> None:
        self.values: dict[str, object] = {}

    def getItem(self, itemKey: str):
        return self.values.get(itemKey)

    def setItem(self, itemKey: str, itemValue, key: str = "set"):
        self.values[itemKey] = itemValue

    def eraseItem(self, itemKey: str, key: str = "eraseItem", default=None):
        self.values.pop(itemKey, None)


def run_smoke(input_path: Path) -> tuple[int, int, int]:
    config = AppConfig.load_default()
    parser = NiceWorkbookParser(config)
    candidates = parser.find_candidates([str(input_path)])
    assert candidates, f"Aucune feuille NICE détectée dans {input_path.name}"
    selected = parser.choose_latest_per_file(candidates)
    intervals, issues = parser.parse(selected)
    assert intervals

    profiles = load_seed_profiles(ROOT / "data" / "agents_seed.json")
    store = MemoryAgentStore(profiles)
    scheduler = PlanningScheduler(config, store)
    result, _ = scheduler.schedule(intervals, issues, [input_path.name])
    assert result.days

    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as handle:
        output = Path(handle.name)
    try:
        ExcelExporter().export(result, str(output))
        assert output.exists() and output.stat().st_size > 0
    finally:
        output.unlink(missing_ok=True)
    uncovered = sum(sum(day.uncovered.values()) for day in result.days)
    return len(intervals), len(result.assignments), uncovered


def test_browser_repository_persists_across_sessions() -> None:
    storage = FakeBrowserStorage()
    first_session: dict[str, object] = {}
    seed = load_seed_profiles(ROOT / "data" / "agents_seed.json")[:3]

    repository = BrowserLocalStorageAgentRepository(storage, first_session)
    repository.ensure_seed(seed)
    loaded = repository.list_profiles()
    assert len(loaded) == 3

    loaded[0].role = "Team Lead"
    loaded[0].excluded = True
    repository.upsert_many([loaded[0]])

    # Simulate a full browser refresh: Streamlit session state is empty, while
    # browser localStorage remains available.
    second_session: dict[str, object] = {}
    reopened = BrowserLocalStorageAgentRepository(storage, second_session)
    reloaded = reopened.list_profiles()
    match = next(item for item in reloaded if item.agent_id == loaded[0].agent_id)
    assert match.role == "Team Lead"
    assert match.excluded is True


def test_json_backup_round_trip() -> None:
    profiles = [
        AgentProfile(
            agent_id="123",
            name="Agent Test",
            role="Open Time",
            excluded=True,
            notes="Test",
        )
    ]
    restored = profiles_from_json_bytes(profiles_to_json(profiles))
    assert len(restored) == 1
    assert restored[0].agent_id == "123"
    assert restored[0].role == "Open Time"
    assert restored[0].excluded is True


def test_deleted_seed_agent_does_not_return() -> None:
    storage = FakeBrowserStorage()
    seed = load_seed_profiles(ROOT / "data" / "agents_seed.json")[:3]
    first_session: dict[str, object] = {}
    repository = BrowserLocalStorageAgentRepository(storage, first_session)
    repository.ensure_seed(seed)

    removed = repository.list_profiles()[0]
    from web_store import profile_key

    repository.delete_many([profile_key(removed.agent_id, removed.name)])
    assert all(item.agent_id != removed.agent_id for item in repository.list_profiles())

    # A new Streamlit session runs ensure_seed again, but a deliberate deletion
    # must remain deleted rather than being restored from the initial directory.
    second_session: dict[str, object] = {}
    reopened = BrowserLocalStorageAgentRepository(storage, second_session)
    reopened.ensure_seed(seed)
    assert all(item.agent_id != removed.agent_id for item in reopened.list_profiles())


def test_delete_all_agents_stays_empty_until_explicit_reset() -> None:
    storage = FakeBrowserStorage()
    seed = load_seed_profiles(ROOT / "data" / "agents_seed.json")[:2]
    repository = BrowserLocalStorageAgentRepository(storage, {})
    repository.ensure_seed(seed)

    from web_store import profile_key

    repository.delete_many([profile_key(item.agent_id, item.name) for item in seed])
    assert repository.list_profiles() == []

    reopened = BrowserLocalStorageAgentRepository(storage, {})
    reopened.ensure_seed(seed)
    assert reopened.list_profiles() == []
    reopened.reset_to_seed(seed)
    assert len(reopened.list_profiles()) == 2


def _build_supervisor_rules_result():
    config = AppConfig.load_default()
    profiles = [
        AgentProfile(str(100 + index), f"Superviseur {index}", role="SUP")
        for index in range(1, 5)
    ] + [
        AgentProfile(str(200 + index), f"Open Time {index}", role="Open Time")
        for index in range(1, 4)
    ]
    intervals = []
    monday = date(2026, 7, 13)
    source_row = 1
    for day_offset in range(5):
        work_date = monday + timedelta(days=day_offset)
        for profile in profiles:
            category = (
                Category.SUPERVISEUR
                if profile.role == "SUP"
                else Category.OPEN_TIME
            )
            intervals.append(
                ActivityInterval(
                    source_file="synthetic.xlsx",
                    source_sheet="Semaine",
                    source_row=source_row,
                    agent_id=profile.agent_id,
                    agent_name=profile.name,
                    work_date=work_date,
                    global_start=8 * 60,
                    global_end=20 * 60,
                    activity=category.value,
                    activity_start=8 * 60,
                    activity_end=20 * 60,
                    category=category,
                )
            )
            source_row += 1
    result, _ = PlanningScheduler(config, MemoryAgentStore(profiles)).schedule(
        intervals,
        [],
        ["synthetic.xlsx"],
    )
    return result


def test_supervisor_rotation_morning_brief_and_qvca() -> None:
    result = _build_supervisor_rules_result()
    special_slot = next(
        slot
        for slot in result.slots
        if slot.start_minute == 19 * 60 and slot.end_minute == 20 * 60
    )
    rotation = Counter()
    for day in result.days:
        assignments = [
            item
            for item in day.assignments
            if item.slot_key == special_slot.key and item.role == "SUP"
        ]
        assert len(assignments) == 2
        rotation.update(item.agent_id for item in assignments)
    assert max(rotation.values()) - min(rotation.values()) <= 1

    activities_by_agent = defaultdict(list)
    for activity in result.planned_activities:
        activities_by_agent[activity.agent_id].append(activity)

    supervisors = [item for item in activities_by_agent if item.startswith("10")]
    assert len(supervisors) == 4
    brief_intervals = set()
    for agent_id in supervisors:
        activities = activities_by_agent[agent_id]
        briefs = [item for item in activities if item.activity_type == "Morning Brief"]
        qvca = [item for item in activities if item.activity_type == "Picking QVCA"]
        assert len(briefs) == 1
        assert briefs[0].work_date.weekday() < 5
        assert 11 * 60 + 30 <= briefs[0].start_minute
        assert briefs[0].end_minute <= 15 * 60 + 30
        assert (briefs[0].work_date, briefs[0].start_minute, briefs[0].end_minute) not in brief_intervals
        brief_intervals.add(
            (briefs[0].work_date, briefs[0].start_minute, briefs[0].end_minute)
        )
        assert sum(item.end_minute - item.start_minute for item in qvca) == 120

    assistance_intervals = defaultdict(list)
    slot_by_key = {slot.key: slot for slot in result.slots}
    for assignment in result.assignments:
        slot = slot_by_key[assignment.slot_key]
        assistance_intervals[(assignment.work_date, assignment.agent_id)].append(
            (slot.start_minute, slot.end_minute)
        )
    for activity in result.planned_activities:
        assert not any(
            start < activity.end_minute and end > activity.start_minute
            for start, end in assistance_intervals[(activity.work_date, activity.agent_id)]
        )


def test_export_contains_supervisor_activities() -> None:
    result = _build_supervisor_rules_result()
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as handle:
        output = Path(handle.name)
    try:
        ExcelExporter().export(result, str(output))
        workbook = load_workbook(output, read_only=True, data_only=True)
        values = {
            cell.value
            for row in workbook["Planning"].iter_rows()
            for cell in row
            if cell.value
        }
        assert "Morning Brief" in values
        assert "Picking QVCA" in values
        recap_headers = [cell.value for cell in next(workbook["Récap Assistances"].iter_rows())]
        assert "Picking QVCA (h)" in recap_headers
        assert "Morning Brief" in recap_headers
    finally:
        output.unlink(missing_ok=True)


if __name__ == "__main__":
    import sys

    for raw in sys.argv[1:]:
        path = Path(raw)
        print(path.name, run_smoke(path))
